"""
Читает xlsx из S3 (tmp/ydisk-raw.xlsx) по чанкам и пишет в БД.
Шаг 0: сначала вызов с { "init": true } — считывает заголовок и кол-во строк, сохраняет мету.
Шаг N: вызов с { "chunk": N, "mode": "replace"|"merge" } — парсит строки N*CHUNK_SIZE..(N+1)*CHUNK_SIZE и пишет в БД.
"""
import json
import os
import io
import re
import boto3
import openpyxl
import psycopg2

CORS_HEADERS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "GET, POST, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type",
}

S3_RAW_KEY  = "tmp/ydisk-raw.xlsx"
S3_META_KEY = "tmp/ydisk-meta.json"
CHUNK_SIZE  = 500

INSERT_MOD = """
    INSERT INTO car_modifications (
        id, generation_id, name, engine, transmission, power,
        body_type, seats, length_mm, width_mm, height_mm, wheelbase_mm,
        track_front_mm, track_rear_mm, curb_weight_kg, wheel_size, ground_clearance_mm,
        trunk_max_l, trunk_min_l, gross_weight_kg, disk_size, clearance_mm,
        track_front_width_mm, track_rear_width_mm, payload_kg, train_weight_kg, axle_load_kg,
        loading_height_mm, cargo_compartment_dims, cargo_volume_m3, bolt_pattern,
        engine_type, engine_volume_cc, power_rpm, torque_nm, intake_type,
        cylinder_layout, cylinder_count, compression_ratio, valves_per_cylinder, turbo_type,
        bore_mm, stroke_mm, engine_model, engine_location, power_kw, torque_rpm,
        intercooler, engine_code, timing_system, fuel_consumption_method,
        gear_count, drive_type, turning_diameter_m,
        fuel_type, max_speed_kmh, acceleration_100, fuel_tank_l, eco_standard,
        fuel_city_l, fuel_highway_l, fuel_mixed_l, range_km, co2_g_km,
        front_brakes, rear_brakes, front_suspension, rear_suspension,
        doors_count, country_of_origin, vehicle_class, steering_position,
        safety_rating, safety_rating_name,
        battery_capacity_kwh, electric_range_km, charge_time_h, battery_type,
        battery_temp_range_c, fast_charge_time_h, fast_charge_desc, charge_connector_type,
        consumption_kwh_per_100km, max_charge_power_kw, battery_available_kwh, charge_cycles
    ) VALUES (
        %s,%s,%s,%s,%s,%s,
        %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
        %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
        %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
        %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
        %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s
    ) ON CONFLICT (id) DO UPDATE SET
        name=EXCLUDED.name, engine=EXCLUDED.engine,
        transmission=EXCLUDED.transmission, power=EXCLUDED.power
"""


def slug(s):
    return re.sub(r"[\s()/\\]+", "-", s.lower()).strip("-")

def make_id(*parts):
    return "__".join(slug(p) for p in parts if p)

def g(row, i):
    v = row[i] if i < len(row) else None
    s = str(v).strip() if v is not None else ""
    if s in ("None", "none", ""):
        return None
    return s.replace(",", ".")

def gn(row, i):
    """Возвращает значение только если оно похожо на число, иначе None."""
    v = g(row, i)
    if v is None:
        return None
    try:
        float(v)
        return v
    except (ValueError, TypeError):
        return None

def gc(row, col_idx, col_name, fallback_i):
    i = col_idx.get(col_name.lower())
    return g(row, i) if i is not None else g(row, fallback_i)

def gcn(row, col_idx, col_name, fallback_i):
    """gc для числовых полей — возвращает None если не число."""
    i = col_idx.get(col_name.lower())
    return gn(row, i) if i is not None else gn(row, fallback_i)


def get_s3():
    return boto3.client(
        "s3",
        endpoint_url="https://bucket.poehali.dev",
        aws_access_key_id=os.environ["AWS_ACCESS_KEY_ID"],
        aws_secret_access_key=os.environ["AWS_SECRET_ACCESS_KEY"],
    )


def init_meta(s3) -> dict:
    """Читает заголовок xlsx и считает строки, сохраняет мету в S3."""
    obj = s3.get_object(Bucket="files", Key=S3_RAW_KEY)
    file_bytes = obj["Body"].read()

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    header_excel_row = 1
    total_data_rows = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_vals = [str(c) if c is not None else "" for c in row]
        if header_row is None:
            if str(row_vals[0]).strip().lower() in ("марка", "brand"):
                header_row = row_vals
                header_excel_row = i
            elif i <= 5:
                header_row = row_vals
                header_excel_row = i
        else:
            if any(c for c in row_vals):
                total_data_rows += 1

    wb.close()

    total_chunks = (total_data_rows + CHUNK_SIZE - 1) // CHUNK_SIZE

    meta = {
        "header": header_row,
        "header_excel_row": header_excel_row,
        "total_rows": total_data_rows,
        "total_chunks": total_chunks,
        "chunk_size": CHUNK_SIZE,
    }
    s3.put_object(
        Bucket="files",
        Key=S3_META_KEY,
        Body=json.dumps(meta, ensure_ascii=False).encode("utf-8"),
        ContentType="application/json",
    )
    return meta


def read_chunk_rows(s3, meta: dict, chunk_index: int) -> list:
    """Читает только нужный диапазон строк из xlsx через read_only openpyxl."""
    obj = s3.get_object(Bucket="files", Key=S3_RAW_KEY)
    file_bytes = obj["Body"].read()

    header_excel_row = meta["header_excel_row"]
    chunk_size = meta["chunk_size"]

    # Excel row numbers (1-based): data starts at header_excel_row+1
    data_start = header_excel_row + 1
    row_from = data_start + chunk_index * chunk_size
    row_to   = row_from + chunk_size - 1

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=row_from, max_row=row_to, values_only=True), start=row_from):
        row_vals = [str(c) if c is not None else "" for c in row]
        if any(c for c in row_vals):
            rows.append(row_vals)

    wb.close()
    return rows


def process_rows(cur, rows, col_idx):
    brands_batch, models_batch, gens_batch, mods_batch = [], [], [], []
    brands_seen, models_seen, gens_seen, mods_seen = set(), set(), set(), set()
    total = skipped = 0

    for row in rows:
        brand_name = gc(row, col_idx, "марка", 0)
        model_name = gc(row, col_idx, "модель", 1)
        gen_name   = gc(row, col_idx, "поколение", 2)
        year_from  = gc(row, col_idx, "год от (поколение)", 3)
        year_to    = gc(row, col_idx, "год до (поколение)", 4)
        series     = gc(row, col_idx, "серия", 5)
        mod_name   = gc(row, col_idx, "модификация", 6)

        if not brand_name or not model_name or not mod_name:
            skipped += 1; continue

        years     = f"{year_from} — {year_to}" if year_to else (year_from or "")
        gen_label = f"{gen_name} {series}".strip() if series else (gen_name or "")
        brand_id  = slug(brand_name)
        model_id  = make_id(brand_id, model_name)
        gen_id    = make_id(model_id, gen_label or mod_name)
        mod_id    = make_id(gen_id, mod_name)

        if mod_id in mods_seen:
            skipped += 1; continue

        if brand_id not in brands_seen:
            brands_batch.append((brand_id, brand_name)); brands_seen.add(brand_id)
        if model_id not in models_seen:
            models_batch.append((model_id, brand_id, model_name)); models_seen.add(model_id)
        if gen_id not in gens_seen:
            gens_batch.append((gen_id, model_id, gen_label or mod_name, years)); gens_seen.add(gen_id)

        def f(name, i): return gc(row, col_idx, name, i)
        def fn(name, i): return gcn(row, col_idx, name, i)

        engine_type      = f("тип двигателя", 32)
        engine_volume_cc = fn("объем двигателя [см3]", 33)
        power_val        = fn("мощность двигателя [л.с.]", 34)
        parts = [p for p in [engine_type, f"{engine_volume_cc} см³" if engine_volume_cc else None, f"{power_val} л.с." if power_val else None] if p]
        engine       = " ".join(parts) if parts else mod_name
        power        = power_val or "—"
        transmission = f("тип кпп", 53) or "—"

        mods_batch.append((
            mod_id, gen_id, mod_name, engine, transmission, power,
            f("тип кузова",7), fn("количество мест",8), fn("длина [мм]",9), fn("ширина [мм]",10),
            fn("высота [мм]",11), fn("колёсная база [мм]",12), fn("колея передняя [мм]",13),
            fn("колея задняя [мм]",14), fn("снаряженная масса [кг]",15), f("размер колёс",16),
            fn("дорожный просвет [мм]",17), fn("объем багажника максимальный [л]",18),
            fn("объем багажника минимальный [л]",19), fn("полная масса [кг]",20),
            f("размер дисков",21), fn("клиренс [мм]",22), fn("ширина передней колеи [мм]",23),
            fn("ширина задней колеи [мм]",24), fn("грузоподъёмность [кг]",25),
            fn("разрешённая масса автопоезда [кг]",26), fn("нагрузка на переднюю/заднюю ось [кг]",27),
            fn("погрузочная высота [мм]",28), f("грузовой отсек (длина x ширина x высота) [мм]",29),
            fn("объём грузового отсека [м3]",30), f("сверловка [мм]",31),
            engine_type, engine_volume_cc, fn("обороты максимальной мощности [об/мин]",35),
            fn("максимальный крутящий момент [н*м]",36), f("тип впуска",37),
            f("расположение цилиндров",38), fn("количество цилиндров",39), fn("степень сжатия",40),
            fn("количество клапанов на цилиндр",41), f("тип наддува",42),
            fn("диаметр цилиндра [мм]",43), fn("ход поршня [мм]",44), f("модель двигателя",45),
            f("расположение двигателя",46), fn("максимальная мощность (квт) [квт]",47),
            fn("обороты максимального крутящего момента [об/мин]",48), f("наличие интеркулера",49),
            f("код двигателя",50), f("грм",51), f("методика расчета расхода",52),
            fn("количество передач",54), f("привод",55), fn("диаметр разворота [м]",56),
            f("марка топлива",57), fn("максимальная скорость [км/ч]",58),
            fn("разгон до 100 км/ч [сек]",59), fn("объём топливного бака [л]",60),
            f("экологический стандарт",61), fn("расход топлива в городе на 100 км [л]",62),
            fn("расход топлива на шоссе на 100 км [л]",63),
            fn("расход топлива в смешанном цикле на 100 км [л]",64),
            fn("запас хода [км]",65), fn("выбросы co2 [г/км]",66),
            f("передние тормоза",67), f("задние тормоза",68),
            f("передняя подвеска",69), f("задняя подвеска",70),
            fn("количество дверей",71), f("страна марки",72), f("класс автомобиля",73),
            f("расположение руля",74), fn("оценка безопасности",75), f("название рейтинга",76),
            fn("емкость батареи [квт⋅ч]",77), fn("запас хода на электричестве [км]",78),
            fn("время зарядки [ч]",79), f("тип батареи",80), f("температурный режим батареи [c]",81),
            fn("время быстрой зарядки [ч]",82), f("описание быстрой зарядки",83),
            f("тип разъема для зарядки",84), fn("расход [квт⋅ч/100 км]",85),
            fn("максимальная мощность зарядки [квт]",86), fn("ёмкость батареи (доступная) [квт⋅ч]",87),
            fn("количество циклов зарядки",88),
        ))
        mods_seen.add(mod_id)
        total += 1

    if brands_batch:
        cur.executemany("INSERT INTO car_brands (id, name) VALUES (%s, %s) ON CONFLICT (id) DO UPDATE SET name=EXCLUDED.name", brands_batch)
    if models_batch:
        cur.executemany("INSERT INTO car_models (id, brand_id, name) VALUES (%s, %s, %s) ON CONFLICT (id) DO UPDATE SET name=EXCLUDED.name", models_batch)
    if gens_batch:
        cur.executemany("INSERT INTO car_generations (id, model_id, name, years) VALUES (%s, %s, %s, %s) ON CONFLICT (id) DO UPDATE SET name=EXCLUDED.name, years=EXCLUDED.years", gens_batch)
    if mods_batch:
        cur.executemany(INSERT_MOD, mods_batch)

    return total, skipped


def handler(event: dict, context) -> dict:
    """Парсит xlsx из S3 по чанкам и пишет в БД. init=true — инициализация меты, chunk=N — загрузка данных."""
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS_HEADERS, "body": ""}

    try:
        body = json.loads(event.get("body") or "{}")
    except Exception:
        return {"statusCode": 400, "headers": CORS_HEADERS, "body": json.dumps({"error": "Неверные параметры"})}

    s3 = get_s3()

    # Режим инициализации: считаем строки, сохраняем мету
    if body.get("init"):
        meta = init_meta(s3)
        return {
            "statusCode": 200,
            "headers": {**CORS_HEADERS, "Content-Type": "application/json"},
            "body": json.dumps({
                "ok": True,
                "total_rows": meta["total_rows"],
                "total_chunks": meta["total_chunks"],
            }),
        }

    # Режим загрузки чанка
    chunk_index = int(body.get("chunk", 0))
    mode        = body.get("mode", "replace")

    meta_obj   = s3.get_object(Bucket="files", Key=S3_META_KEY)
    meta       = json.loads(meta_obj["Body"].read().decode("utf-8"))
    header_row = meta["header"]
    total_chunks = meta["total_chunks"]
    total_rows   = meta["total_rows"]
    col_idx      = {str(h).strip().lower(): i for i, h in enumerate(header_row)}

    rows_chunk = read_chunk_rows(s3, meta, chunk_index)

    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    cur  = conn.cursor()

    if chunk_index == 0 and mode == "replace":
        cur.execute("TRUNCATE car_modifications, car_generations, car_models, car_brands RESTART IDENTITY CASCADE")

    ins, skp = process_rows(cur, rows_chunk, col_idx)
    conn.commit()
    cur.close()
    conn.close()

    return {
        "statusCode": 200,
        "headers": {**CORS_HEADERS, "Content-Type": "application/json"},
        "body": json.dumps({
            "inserted": ins, "skipped": skp,
            "total_rows": total_rows, "total_chunks": total_chunks,
            "chunk": chunk_index, "done": chunk_index >= total_chunks - 1,
        }),
    }